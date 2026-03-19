from typing import Optional, List, Dict, Any, Tuple
from datetime import datetime, timedelta
from io import BytesIO
from collections import Counter
from sqlalchemy.orm import Session

# Models
from src.models.admin import (
    Staff, StaffCreate, StaffUpdate, StaffRole, StaffStatus,
    Bed, InventoryItem, ActivityLog, AdminModel
)
from src.models.patients import Consultation, Patient

# For PDF generation
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.barcharts import VerticalBarChart

# For Excel
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


class AdminService:
    
    # ========== Staff Management ==========
    
    @staticmethod
    def get_all_staff(db: Session, role: Optional[str] = None, department: Optional[str] = None) -> List[Staff]:
        return AdminModel.get_all_staff(role=role, department=department, db=db)

    @staticmethod
    def get_staff_by_id(db: Session, staff_id: int) -> Optional[Staff]:
        return AdminModel.get_staff_by_id(staff_id, db=db)

    @staticmethod
    def get_staff_by_staff_id(db: Session, staff_id_str: str) -> Optional[Staff]:
        return AdminModel.get_staff_by_staff_id(staff_id_str, db=db)

    @staticmethod
    def create_staff(db: Session, staff_data: StaffCreate) -> Staff:
        return AdminModel.create_staff(staff_data, db=db)

    @staticmethod
    def update_staff(db: Session, staff_id: int, staff_update: StaffUpdate) -> Optional[Staff]:
        return AdminModel.update_staff(staff_id, staff_update, db=db)

    @staticmethod
    def delete_staff(db: Session, staff_id: int) -> bool:
        return AdminModel.delete_staff(staff_id, db=db)

    # ========== Bed Management ==========
    
    @staticmethod
    def get_bed_status(db: Session) -> Dict:
        return AdminModel.get_bed_status(db=db)

    @staticmethod
    def assign_bed(db: Session, bed_id: int, patient_id: int) -> Optional[Bed]:
        return AdminModel.assign_bed(bed_id, patient_id, db=db)

    @staticmethod
    def release_bed(db: Session, bed_id: int) -> Optional[Bed]:
        return AdminModel.release_bed(bed_id, db=db)

    # ========== Inventory Management ==========
    
    @staticmethod
    def get_inventory_status(db: Session) -> Dict:
        return AdminModel.get_inventory_status(db=db)

    @staticmethod
    def update_inventory(db: Session, item_id: int, quantity_change: int) -> Optional[InventoryItem]:
        return AdminModel.update_inventory(item_id, quantity_change, db=db)

    # ========== Activity Log ==========
    
    @staticmethod
    def log_activity(db: Session, user_id: int, user_name: str, role: str, action: str, details: str):
        AdminModel.log_activity(
            user_id=user_id, 
            user_name=user_name, 
            action=action, 
            details=details, 
            role=role, 
            db=db
        )

    @staticmethod
    def get_recent_activity(db: Session, limit: int = 10) -> List[ActivityLog]:
        return AdminModel.get_activity_logs(limit=limit, db=db)

    # ========== Dashboard Statistics ==========
    
    @staticmethod
    def get_dashboard_stats(db: Session) -> Dict:
        staff_stats = AdminModel.get_staff_statistics(db)
        bed_stats = AdminModel.get_bed_status(db)
        inv_stats = AdminModel.get_inventory_status(db)
        logs = AdminModel.get_activity_logs(limit=5, db=db)
        recent_activity = [
            {
                "user": log.user_name,
                "action": log.action,
                "time": log.timestamp.isoformat(),
                "details": log.details
            } for log in logs
        ]
        return {
            "staff": staff_stats,
            "beds": bed_stats,
            "inventory": inv_stats,
            "recent_activity": recent_activity,
            "server_time": datetime.now().isoformat()
        }

    # ========== Report Generation ==========

    @staticmethod
    async def generate_report(
        format: str,
        department: Optional[str] = None,
        from_date: Optional[str] = None,
        to_date: Optional[str] = None,
        status: Optional[str] = None,
        include_charts: bool = False,
        db: Session = None,
    ) -> Tuple[bytes, str, str]:
        result = await AdminService._get_filtered_consultations_db(
            department, from_date, to_date, status, db
        )
        records = result["records"]
        stats = result["stats"]

        if format.lower() == "pdf":
            if include_charts:
                return await AdminService._generate_pdf_with_charts(records, stats, db)
            else:
                return await AdminService._generate_pdf(records)
        elif format.lower() == "excel":
            return await AdminService._generate_excel(records)
        else:
            raise ValueError("Unsupported format. Use 'pdf' or 'excel'.")

    @staticmethod
    async def _get_filtered_consultations_db(
        department: Optional[str],
        from_date: Optional[str],
        to_date: Optional[str],
        status: Optional[str],
        db: Session
    ) -> Dict:
        query = db.query(Consultation).join(Patient, Consultation.patient_id == Patient.id)

        if department:
            query = query.filter(Consultation.department == department)
        if status:
            query = query.filter(Consultation.status == status)
        if from_date:
            from_dt = datetime.fromisoformat(from_date)
            query = query.filter(Consultation.created_at >= from_dt)
        if to_date:
            to_dt = datetime.fromisoformat(to_date) + timedelta(days=1)
            query = query.filter(Consultation.created_at < to_dt)

        consultations = query.all()

        records = []
        for c in consultations:
            patient = db.query(Patient).filter(Patient.id == c.patient_id).first()
            doctor = db.query(Staff).filter(Staff.staff_id == c.doctor_id).first() if c.doctor_id else None
            records.append({
                "id": c.id,
                "patient_name": patient.name if patient else "Unknown",
                "doctor_name": doctor.name if doctor else "Unknown",
                "department": c.department,
                "created_at": c.created_at.isoformat() if c.created_at else None,
                "status": c.status,
                "condition": c.condition,
                "priority": c.priority,
            })

        total_patients = db.query(Patient).count()
        total_consultations = len(consultations)
        wait_times = []
        for c in consultations:
            if c.started_at and c.created_at:
                wait = (c.started_at - c.created_at).seconds // 60
                if wait > 0:
                    wait_times.append(wait)
        avg_wait_time = sum(wait_times) // len(wait_times) if wait_times else 0
        treated_today = db.query(Consultation).filter(
            Consultation.status == "treated",
            Consultation.completed_at >= datetime.now().replace(hour=0, minute=0, second=0)
        ).count()

        stats = {
            "totalPatients": total_patients,
            "totalConsultations": total_consultations,
            "avgWaitTime": avg_wait_time,
            "treatedToday": treated_today,
        }
        return {"records": records, "stats": stats}

    @staticmethod
    async def _generate_pdf(records: List[Dict]) -> Tuple[bytes, str, str]:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("Mashar Hospital - Consultation Report", styles['Title']))
        elements.append(Spacer(1, 0.2 * inch))

        data = [["Patient", "Doctor", "Department", "Date", "Status", "Condition"]]
        for r in records:
            date_str = datetime.fromisoformat(r["created_at"]).strftime("%Y-%m-%d %H:%M") if r["created_at"] else ""
            data.append([
                r["patient_name"],
                r["doctor_name"],
                r["department"],
                date_str,
                r["status"],
                r["condition"],
            ])

        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue(), "application/pdf", "report.pdf"

    @staticmethod
    async def _generate_pdf_with_charts(records: List[Dict], stats: Dict, db: Session) -> Tuple[bytes, str, str]:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("Mashar Hospital - Consultation Report (with Charts)", styles['Title']))
        elements.append(Spacer(1, 0.2 * inch))

        stats_text = f"""
        Total Patients: {stats['totalPatients']} | Total Consultations: {stats['totalConsultations']} | Avg Wait Time: {stats['avgWaitTime']} min | Treated Today: {stats['treatedToday']}
        """
        elements.append(Paragraph(stats_text, styles['Normal']))
        elements.append(Spacer(1, 0.2 * inch))

        dept_counts = Counter(r['department'] for r in records if r['department'])
        if dept_counts:
            categories = list(dept_counts.keys())
            values = list(dept_counts.values())
            chart_data = [values]

            drawing = Drawing(400, 200)
            bc = VerticalBarChart()
            bc.x = 50
            bc.y = 50
            bc.width = 300
            bc.height = 150
            bc.data = chart_data
            bc.strokeColor = colors.black
            bc.valueAxis.valueMin = 0
            bc.valueAxis.valueMax = max(values) + 1 if values else 1
            bc.valueAxis.valueStep = 1
            bc.categoryAxis.categoryNames = categories
            bc.categoryAxis.labels.boxAnchor = 'ne'
            bc.categoryAxis.labels.dx = 8
            bc.categoryAxis.labels.dy = -2
            bc.categoryAxis.labels.angle = 45
            bc.bars[0].fillColor = colors.HexColor('#0a8a7c')
            bc.bars[0].strokeColor = colors.black

            drawing.add(bc)
            elements.append(drawing)
            elements.append(Spacer(1, 0.3 * inch))
        else:
            elements.append(Paragraph("No department data available for chart.", styles['Normal']))
            elements.append(Spacer(1, 0.2 * inch))

        data = [["Patient", "Doctor", "Department", "Date", "Status", "Condition"]]
        for r in records:
            date_str = datetime.fromisoformat(r["created_at"]).strftime("%Y-%m-%d %H:%M") if r["created_at"] else ""
            data.append([
                r["patient_name"],
                r["doctor_name"],
                r["department"],
                date_str,
                r["status"],
                r["condition"],
            ])

        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue(), "application/pdf", "report_with_charts.pdf"

    @staticmethod
    async def _generate_excel(records: List[Dict]) -> Tuple[bytes, str, str]:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Consultations"

        headers = ["Patient", "Doctor", "Department", "Date", "Status", "Condition"]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0A8A7C", end_color="0A8A7C", fill_type="solid")
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for r in records:
            date_str = datetime.fromisoformat(r["created_at"]).strftime("%Y-%m-%d %H:%M") if r["created_at"] else ""
            ws.append([
                r["patient_name"],
                r["doctor_name"],
                r["department"],
                date_str,
                r["status"],
                r["condition"],
            ])

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[col_letter].width = adjusted_width

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "report.xlsx"

    # ===== Personnel Export with Charts =====
    @staticmethod
    async def generate_personnel_report(
        format: str,
        db: Session,
        include_charts: bool = False
    ) -> Tuple[bytes, str, str]:
        staff = db.query(Staff).all()
        patients = db.query(Patient).all()

        if format.lower() == "pdf":
            if include_charts:
                return await AdminService._generate_personnel_pdf_with_charts(staff, patients, db)
            else:
                return await AdminService._generate_personnel_pdf(staff, patients)
        elif format.lower() == "excel":
            return await AdminService._generate_personnel_excel(staff, patients)
        else:
            raise ValueError("Unsupported format. Use 'pdf' or 'excel'.")

    @staticmethod
    async def _generate_personnel_pdf(staff: List[Staff], patients: List[Patient]) -> Tuple[bytes, str, str]:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("Mashar Hospital - Personnel Records", styles['Title']))
        elements.append(Spacer(1, 0.2 * inch))

        # Staff table
        elements.append(Paragraph("Staff List", styles['Heading2']))
        staff_data = [["ID", "Name", "Role", "Department", "Phone"]]
        for s in staff:
            staff_data.append([
                s.staff_id,
                s.name,
                s.role.value if hasattr(s.role, 'value') else str(s.role),
                s.department or "",
                s.phone or ""
            ])
        staff_table = Table(staff_data)
        staff_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(staff_table)
        elements.append(Spacer(1, 0.3 * inch))

        # Patients table
        elements.append(Paragraph("Patient List", styles['Heading2']))
        patient_data = [["ID", "Name", "Phone", "Total Visits", "Last Visit"]]
        for p in patients:
            total_visits = getattr(p, 'total_visits', 0)
            last_visit = getattr(p, 'last_visit', None)
            patient_data.append([
                str(p.id),
                p.name,
                p.phone or "",
                str(total_visits),
                last_visit.strftime("%Y-%m-%d") if last_visit else ""
            ])
        patient_table = Table(patient_data)
        patient_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(patient_table)

        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue(), "application/pdf", "personnel_report.pdf"

    @staticmethod
    async def _generate_personnel_excel(staff: List[Staff], patients: List[Patient]) -> Tuple[bytes, str, str]:
        wb = openpyxl.Workbook()

        # Staff sheet
        ws_staff = wb.active
        ws_staff.title = "Staff"
        staff_headers = ["ID", "Name", "Role", "Department", "Phone"]
        ws_staff.append(staff_headers)
        for s in staff:
            ws_staff.append([
                s.staff_id,
                s.name,
                s.role.value if hasattr(s.role, 'value') else str(s.role),
                s.department or "",
                s.phone or ""
            ])

        # Patients sheet
        ws_patients = wb.create_sheet("Patients")
        patient_headers = ["ID", "Name", "Phone", "Total Visits", "Last Visit"]
        ws_patients.append(patient_headers)
        for p in patients:
            total_visits = getattr(p, 'total_visits', 0)
            last_visit = getattr(p, 'last_visit', None)
            ws_patients.append([
                p.id,
                p.name,
                p.phone or "",
                total_visits,
                last_visit.strftime("%Y-%m-%d") if last_visit else ""
            ])

        for ws in [ws_staff, ws_patients]:
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[col_letter].width = adjusted_width

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "personnel_report.xlsx"

    @staticmethod
    async def _generate_personnel_pdf_with_charts(
        staff: List[Staff],
        patients: List[Patient],
        db: Session
    ) -> Tuple[bytes, str, str]:
        """Generate a PDF with staff and patient tables plus charts (safe attribute access)."""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("Mashar Hospital - Personnel Records (with Charts)", styles['Title']))
        elements.append(Spacer(1, 0.2 * inch))

        # --- Staff Chart: Count by Role ---
        role_counts = Counter(s.role.value if hasattr(s.role, 'value') else str(s.role) for s in staff)
        if role_counts:
            elements.append(Paragraph("Staff Distribution by Role", styles['Heading2']))
            categories = list(role_counts.keys())
            values = list(role_counts.values())
            chart_data = [values]

            drawing = Drawing(400, 200)
            bc = VerticalBarChart()
            bc.x = 50
            bc.y = 50
            bc.width = 300
            bc.height = 150
            bc.data = chart_data
            bc.strokeColor = colors.black
            bc.valueAxis.valueMin = 0
            bc.valueAxis.valueMax = max(values) + 1 if values else 1
            bc.valueAxis.valueStep = 1
            bc.categoryAxis.categoryNames = categories
            bc.categoryAxis.labels.boxAnchor = 'ne'
            bc.categoryAxis.labels.dx = 8
            bc.categoryAxis.labels.dy = -2
            bc.categoryAxis.labels.angle = 45
            bc.bars[0].fillColor = colors.HexColor('#0a8a7c')
            bc.bars[0].strokeColor = colors.black

            drawing.add(bc)
            elements.append(drawing)
            elements.append(Spacer(1, 0.3 * inch))
        else:
            elements.append(Paragraph("No staff data for chart.", styles['Normal']))
            elements.append(Spacer(1, 0.2 * inch))

        # --- Patients Chart: Count by Department (safe) ---
        dept_counts = Counter()
        for p in patients:
            dept = getattr(p, 'department', None)
            if dept:
                dept_counts[dept] += 1
        if dept_counts:
            elements.append(Paragraph("Patients by Department", styles['Heading2']))
            categories = list(dept_counts.keys())
            values = list(dept_counts.values())
            chart_data = [values]

            drawing = Drawing(400, 200)
            bc = VerticalBarChart()
            bc.x = 50
            bc.y = 50
            bc.width = 300
            bc.height = 150
            bc.data = chart_data
            bc.strokeColor = colors.black
            bc.valueAxis.valueMin = 0
            bc.valueAxis.valueMax = max(values) + 1 if values else 1
            bc.valueAxis.valueStep = 1
            bc.categoryAxis.categoryNames = categories
            bc.categoryAxis.labels.boxAnchor = 'ne'
            bc.categoryAxis.labels.dx = 8
            bc.categoryAxis.labels.dy = -2
            bc.categoryAxis.labels.angle = 45
            bc.bars[0].fillColor = colors.HexColor('#12b09f')
            bc.bars[0].strokeColor = colors.black

            drawing.add(bc)
            elements.append(drawing)
            elements.append(Spacer(1, 0.3 * inch))
        else:
            elements.append(Paragraph("No patient department data for chart.", styles['Normal']))
            elements.append(Spacer(1, 0.2 * inch))

        # --- Staff Table ---
        elements.append(Paragraph("Staff List", styles['Heading2']))
        staff_data = [["ID", "Name", "Role", "Department", "Phone"]]
        for s in staff:
            staff_data.append([
                s.staff_id,
                s.name,
                s.role.value if hasattr(s.role, 'value') else str(s.role),
                s.department or "",
                s.phone or ""
            ])
        staff_table = Table(staff_data)
        staff_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(staff_table)
        elements.append(Spacer(1, 0.3 * inch))

        # --- Patients Table ---
        elements.append(Paragraph("Patient List", styles['Heading2']))
        patient_data = [["ID", "Name", "Phone", "Total Visits", "Last Visit"]]
        for p in patients:
            total_visits = getattr(p, 'total_visits', 0)
            last_visit = getattr(p, 'last_visit', None)
            patient_data.append([
                str(p.id),
                p.name,
                p.phone or "",
                str(total_visits),
                last_visit.strftime("%Y-%m-%d") if last_visit else ""
            ])
        patient_table = Table(patient_data)
        patient_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(patient_table)

        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue(), "application/pdf", "personnel_report_with_charts.pdf"